#!/usr/bin/env python3
"""Build Albirex Niigata offseason workbooks from CSV source data.

This script is intended to run in GitHub Actions only. It creates Excel
workbooks from CSV source data; Codex must not execute it locally.
"""
from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.pagebreak import Break

ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT / "data"
EXPORTS_ROOT = ROOT / "exports"

PLAYER_HEADERS = ["背番号", "ポジション", "選手名", "その他備考"]
EVENT_HEADERS = ["選手名", "動向", "発表日", "区分", "移籍先元", "ステータス", "出典URL", "その他備考"]
MOBILE_HEADERS = ["背番号", "ポジション", "選手名", "動向", "発表日", "その他備考", "区分", "移籍先/元", "ステータス", "出典URL"]
LOG_HEADERS = ["確認日", "対象", "確認内容", "結果", "その他備考"]

CATEGORY_MASTER = [
    ("契約更新", "残留系", "薄い緑"),
    ("期限付移籍in延長", "残留系", "薄い緑"),
    ("復帰", "加入系", "薄い青"),
    ("移籍in", "加入系", "薄い青"),
    ("期限付移籍in", "加入系", "薄い青"),
    ("期限付移籍out終了", "加入系", "薄い青"),
    ("移籍out", "退団・離脱系", "薄い橙"),
    ("期限付移籍out", "退団・離脱系", "薄い橙"),
    ("期限付移籍in終了", "退団・離脱系", "薄い橙"),
    ("契約満了", "退団・離脱系", "薄い橙"),
    ("引退", "退団・離脱系", "薄い橙"),
    ("期限付移籍out延長", "その他系", "薄い紫"),
    ("未発表", "未発表", "薄い黄"),
]

COLOR_MAP = {
    "薄い緑": "E2F0D9",
    "薄い青": "DDEBF7",
    "薄い橙": "FCE4D6",
    "薄い紫": "E4DFEC",
    "薄い黄": "FFF2CC",
}

WORKBOOKS = {
    "men": "albirex_niigata_men_2026_27_offseason_print_mobile.xlsx",
    "ladies": "albirex_niigata_ladies_2026_27_offseason_print_mobile.xlsx",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def append_table(ws, headers: list[str], rows: Iterable[Iterable[str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    ws.auto_filter.ref = ws.dimensions


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def text_width_units(value: str) -> float:
    """Estimate wrapped-cell width using wider units for Japanese text."""
    width = 0.0
    for character in value:
        if character == "\n":
            width += 100.0
        elif ord(character) < 128:
            width += 0.55
        else:
            width += 1.0
    return width


def estimate_wrapped_lines(value: str, column_width: float) -> int:
    """Estimate display lines for wrapped text in an Excel column."""
    text = str(value or "")
    if not text:
        return 1
    usable_width = max(column_width - 1.5, 1.0)
    lines = 0
    for paragraph in text.splitlines() or [""]:
        lines += max(1, int((text_width_units(paragraph) + usable_width - 0.01) // usable_width))
    return lines


def mobile_row_height(movement: str, note: str) -> float:
    """Return a conservative row height so wrapped D/F text is not clipped."""
    movement_lines = estimate_wrapped_lines(movement, 16)
    note_lines = estimate_wrapped_lines(note, 30)
    wrapped_lines = max(movement_lines, note_lines)
    return max(34, 18 * wrapped_lines + 14)


def last_data_row(ws, min_col: int = 1, max_col: int = 6) -> int:
    """Return the last row containing actual print data in the target columns."""
    for row_number in range(ws.max_row, 0, -1):
        for column_number in range(min_col, max_col + 1):
            value = ws.cell(row=row_number, column=column_number).value
            if value is not None and str(value).strip():
                return row_number
    return 1


def apply_mobile_page_breaks(ws, max_row: int, min_tail_rows: int = 3) -> None:
    """Keep tall wrapped rows away from the printable page bottom.

    LibreOffice PDF export can be unstable when the final page contains only
    one data row. Prefer an earlier page break so the final PDF page carries a
    small block of rows instead of a single trailing player.
    """
    printable_page_height = 630
    header_height = ws.row_dimensions[1].height or 26
    current_page_height = header_height
    break_ids: list[int] = []
    for row_number in range(2, max_row + 1):
        row_height = ws.row_dimensions[row_number].height or 34
        if current_page_height > header_height and current_page_height + row_height > printable_page_height:
            break_ids.append(row_number - 1)
            current_page_height = header_height + row_height
        else:
            current_page_height += row_height

    if break_ids:
        tail_rows = max_row - break_ids[-1]
        if 0 < tail_rows < min_tail_rows:
            previous_break = break_ids[-2] if len(break_ids) > 1 else 1
            adjusted_break = max(previous_break + 1, max_row - min_tail_rows)
            if adjusted_break < break_ids[-1]:
                break_ids[-1] = adjusted_break

    for break_id in break_ids:
        if 1 <= break_id < max_row:
            ws.row_breaks.append(Break(id=break_id))


def parse_event_date(value: str) -> date | None:
    try:
        return date.fromisoformat(value.strip()) if value.strip() else None
    except ValueError:
        return None


def pick_latest_events(events: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    """Pick one event per player.

    If the same player has multiple events, priority is:
    1. newest 発表日;
    2. if 発表日 is blank or tied, the later row in the CSV.
    """
    picked: dict[str, tuple[date | None, int, dict[str, str]]] = {}
    for index, event in enumerate(events):
        name = event.get("選手名", "").strip()
        if not name:
            continue
        current_key = (parse_event_date(event.get("発表日", "")), index)
        previous = picked.get(name)
        if previous is None:
            picked[name] = (*current_key, event)
            continue
        previous_date, previous_index, _previous_event = previous
        current_date, current_index = current_key
        if current_date and (previous_date is None or current_date > previous_date):
            picked[name] = (*current_key, event)
        elif current_date == previous_date and current_index > previous_index:
            picked[name] = (*current_key, event)
    return {name: item[2] for name, item in picked.items()}


def build_mobile_rows(players: list[dict[str, str]], events: list[dict[str, str]]) -> list[list[str]]:
    events_by_name = pick_latest_events(events)
    rows = []
    for player in players:
        name = player.get("選手名", "")
        event = events_by_name.get(name)
        if event:
            movement = event.get("動向", "") or "未発表"
            announcement_date = event.get("発表日", "")
            note = event.get("その他備考", "")
            category = event.get("区分", "") or "未発表"
            transfer_from_to = event.get("移籍先元", "")
            status = event.get("ステータス", "") or "未発表"
            source_url = event.get("出典URL", "")
        else:
            movement = "未発表"
            announcement_date = ""
            note = player.get("その他備考", "")
            category = "未発表"
            transfer_from_to = ""
            status = "未発表"
            source_url = ""
        rows.append([
            player.get("背番号", ""),
            player.get("ポジション", ""),
            name,
            movement,
            announcement_date,
            note,
            category,
            transfer_from_to,
            status,
            source_url,
        ])
    return rows


def setup_mobile_sheet(ws, rows: list[list[str]]) -> None:
    append_table(ws, MOBILE_HEADERS, rows)
    max_row = last_data_row(ws)
    ws.freeze_panes = "A2"
    ws.print_title_rows = "1:1"
    ws.print_area = f"A1:F{max_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.row_dimensions[1].height = 26
    for row_number in range(2, max_row + 1):
        movement = str(ws.cell(row=row_number, column=4).value or "")
        note = str(ws.cell(row=row_number, column=6).value or "")
        ws.row_dimensions[row_number].height = mobile_row_height(movement, note)
    apply_mobile_page_breaks(ws, max_row)
    set_widths(ws, {"A": 7, "B": 11, "C": 17, "D": 16, "E": 12, "F": 30, "G": 18, "H": 24, "I": 16, "J": 42})
    for column in ("G", "H", "I", "J"):
        ws.column_dimensions[column].hidden = True
    for row_number in range(2, max_row + 1):
        url_cell = ws.cell(row=row_number, column=10)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    apply_conditional_formatting(ws, max_row)


def apply_conditional_formatting(ws, max_row: int) -> None:
    if max_row < 2:
        return
    rules_by_color: dict[str, list[str]] = {}
    for category, _group, color_name in CATEGORY_MASTER:
        rules_by_color.setdefault(COLOR_MAP[color_name], []).append(category)
    for color, categories in rules_by_color.items():
        quoted = ",".join(f'"{category}"' for category in categories)
        formula = f'OR(ISNUMBER(MATCH($D2,{{{quoted}}},0)),ISNUMBER(MATCH($G2,{{{quoted}}},0)))'
        fill = PatternFill("solid", fgColor=color)
        ws.conditional_formatting.add(f"A2:J{max_row}", FormulaRule(formula=[formula], fill=fill))


def setup_simple_sheet(ws, headers: list[str], rows: Iterable[Iterable[str]], widths: dict[str, float] | None = None) -> None:
    append_table(ws, headers, rows)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        set_widths(ws, widths)


def build_workbook(team_key: str, output_name: str) -> None:
    players = read_csv(DATA_ROOT / team_key / "players.csv")
    events = read_csv(DATA_ROOT / team_key / "official_events.csv")

    wb = Workbook()
    wb.remove(wb.active)

    mobile_ws = wb.create_sheet("スマホ表示")
    setup_mobile_sheet(mobile_ws, build_mobile_rows(players, events))

    players_ws = wb.create_sheet("選手マスタ")
    setup_simple_sheet(players_ws, PLAYER_HEADERS, ([row.get(header, "") for header in PLAYER_HEADERS] for row in players), {"A": 8, "B": 12, "C": 22, "D": 34})

    events_ws = wb.create_sheet("オフ動向_入力")
    setup_simple_sheet(events_ws, EVENT_HEADERS, ([row.get(header, "") for header in EVENT_HEADERS] for row in events), {"A": 22, "B": 18, "C": 12, "D": 18, "E": 24, "F": 16, "G": 44, "H": 34})

    category_ws = wb.create_sheet("区分マスタ")
    setup_simple_sheet(category_ws, ["区分", "分類", "色"], CATEGORY_MASTER, {"A": 20, "B": 14, "C": 12})

    log_ws = wb.create_sheet("確認ログ")
    setup_simple_sheet(log_ws, LOG_HEADERS, [], {"A": 14, "B": 24, "C": 36, "D": 18, "E": 30})

    memo_ws = wb.create_sheet("メモ")
    memo_rows = [
        ["このExcelファイルは直接編集せず、dataフォルダ内のCSVを更新してGitHub Actionsで再生成する。"],
        ["スマホ表示シートの印刷範囲は A:F とする。"],
        ["G:J列は管理用情報であり、印刷対象外とする。"],
    ]
    setup_simple_sheet(memo_ws, ["運用上の注意点"], memo_rows, {"A": 90})

    EXPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    wb.save(EXPORTS_ROOT / output_name)


def main() -> None:
    for team_key, output_name in WORKBOOKS.items():
        build_workbook(team_key, output_name)


if __name__ == "__main__":
    main()
