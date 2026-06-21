#!/usr/bin/env python3
"""Prepare a temporary workbook for PDF export.

This helper is intended to run in GitHub Actions from scripts/export_pdf.sh.
It keeps only the スマホ表示 sheet so management sheets are not exported to PDF.
Codex must not execute this helper locally because it reads/writes .xlsx files.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

MOBILE_SHEET_NAME = "スマホ表示"


def prepare_pdf_workbook(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    if MOBILE_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"{MOBILE_SHEET_NAME} sheet is missing: {input_path}")

    for sheet_name in list(wb.sheetnames):
        if sheet_name != MOBILE_SHEET_NAME:
            del wb[sheet_name]

    ws = wb[MOBILE_SHEET_NAME]
    wb.active = wb.sheetnames.index(MOBILE_SHEET_NAME)
    ws.print_area = f"A1:F{max(ws.max_row, 1)}"
    ws.print_title_rows = "1:1"
    ws.sheet_view.showGridLines = False
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: prepare_pdf_workbook.py INPUT_XLSX OUTPUT_XLSX")
    prepare_pdf_workbook(Path(sys.argv[1]), Path(sys.argv[2]))


if __name__ == "__main__":
    main()
