from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
TODAY = date.today().isoformat()

CATEGORY_ROWS = [
    ("契約更新", "契約更新", 10, "緑", "発表日: {announcement_date}"),
    ("完全移籍加入", "完全加入", 20, "青", "{from_club}から加入"),
    ("完全移籍退団", "完全移籍", 30, "橙", "{to_club}へ移籍"),
    ("レンタル移籍加入", "レンタル加入", 40, "青", "{from_club}から期限付き"),
    ("レンタル移籍延長", "レンタル延長", 50, "青", "{period}"),
    ("レンタル移籍退団", "レンタル移籍", 60, "橙", "{to_club}へ期限付き"),
    ("レンタル復帰", "復帰", 70, "青", "{from_club}から復帰"),
    ("契約満了", "契約満了", 80, "橙", "退団"),
    ("現役引退", "引退", 90, "橙", "現役引退"),
    ("レンタル中", "レンタル中", 95, "黄", "{to_club}へ期限付き移籍中"),
    ("未発表", "未発表", 99, "黄", "公式発表未確認"),
    ("公式発表未確認", "未発表", 99, "黄", "公式発表未確認"),
]

COLOR_FILLS = {
    "緑": PatternFill("solid", fgColor="D9EAD3"),
    "青": PatternFill("solid", fgColor="D9EAF7"),
    "橙": PatternFill("solid", fgColor="FCE4D6"),
    "黄": PatternFill("solid", fgColor="FFF2CC"),
    "赤": PatternFill("solid", fgColor="F4CCCC"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def value(row: dict[str, Any], key: str, default: str = "") -> str:
    v = row.get(key, default)
    return "" if v is None else str(v)


def sheet_rows(ws, header_row: int = 1) -> list[dict[str, Any]]:
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        item = {str(headers[c - 1]): ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        if any(v not in (None, "") for v in item.values()):
            rows.append(item)
    return rows


def slug(text: str) -> str:
    table = str.maketrans({" ": "_", "　": "_", "/": "_", "・": "_", "-": "_"})
    return text.translate(table).lower()


def normalize_category(raw: str) -> str:
    if raw in {"完全移籍", "完全移籍（退団）"}:
        return "完全移籍退団"
    if raw == "期限付き移籍より復帰":
        return "レンタル復帰"
    return raw or "未発表"


def build_note(category: str, date_text: str, from_club: str, to_club: str, period: str, notes: str) -> str:
    parts = []
    if period:
        parts.append(period)
    club = to_club or from_club
    if club and club not in "からへ":
        parts.append(club)
    if notes:
        parts.append(notes)
    if not parts and date_text:
        parts.append(f"{date_text}発表")
    if category in {"契約満了", "現役引退"} and not parts:
        parts.append("退団")
    return "、".join(parts)



def extract_standard(path: Path):
    wb = load_workbook(path)
    players = [
        {
            "player_id": value(row, "player_id"),
            "team": value(row, "team"),
            "season": value(row, "season"),
            "number": value(row, "number"),
            "position": value(row, "position"),
            "name": value(row, "name"),
            "player_page_url": value(row, "player_page_url"),
            "base_status": value(row, "base_status"),
            "notes": value(row, "notes"),
        }
        for row in sheet_rows(wb["選手マスタ"], 1)
        if value(row, "player_id")
    ]
    events = [
        {
            "event_id": value(row, "event_id"),
            "player_id": value(row, "player_id"),
            "announcement_date": value(row, "announcement_date"),
            "category": value(row, "category"),
            "movement_summary": value(row, "movement_summary"),
            "from_club": value(row, "from_club"),
            "to_club": value(row, "to_club"),
            "period": value(row, "period"),
            "official_url": value(row, "official_url"),
            "source_title": value(row, "source_title"),
            "verification_status": value(row, "verification_status"),
            "checked_at": value(row, "checked_at"),
            "smartphone_note": value(row, "smartphone_note"),
            "internal_notes": value(row, "internal_notes"),
        }
        for row in sheet_rows(wb["オフ動向_入力"], 1)
        if value(row, "event_id")
    ]
    return players, events

def extract_men(path: Path):
    wb = load_workbook(path)
    players = []
    player_seen = set()
    ws_players = wb["選手一覧"]
    for row in sheet_rows(ws_players, 4):
        name = value(row, "氏名")
        if not name:
            continue
        player_id = f"men_2026_{value(row, '背番号', '未定')}_{slug(name)}"
        if player_id in player_seen:
            continue
        player_seen.add(player_id)
        players.append({
            "player_id": player_id,
            "team": "men",
            "season": "2026/27",
            "number": value(row, "背番号"),
            "position": value(row, "ポジション"),
            "name": name,
            "player_page_url": value(row, "出典URL"),
            "base_status": value(row, "レンタル区分") or "登録選手",
            "notes": value(row, "備考"),
        })
    by_name = {p["name"]: p for p in players}
    events = []
    ws_events = wb["オフ動向"]
    for i, row in enumerate(sheet_rows(ws_events, 1), start=1):
        name = value(row, "氏名")
        if not name:
            continue
        p = by_name.get(name)
        if not p:
            number = value(row, "背番号") or "未定"
            p = {"player_id": f"men_2026_{number}_{slug(name)}", "team": "men", "season": "2026/27", "number": number, "position": value(row, "ポジション"), "name": name, "player_page_url": "", "base_status": "退団/移籍", "notes": ""}
            players.append(p); by_name[name] = p
        category = normalize_category(value(row, "種別"))
        from_club = value(row, "移籍先/元") if "から" in value(row, "移籍先/元") else ""
        to_club = value(row, "移籍先/元") if "へ" in value(row, "移籍先/元") else ""
        period = value(row, "期間")
        events.append({
            "event_id": f"men_{value(row, '発表日')}_{slug(name)}_{i}",
            "player_id": p["player_id"],
            "announcement_date": value(row, "発表日"),
            "category": category,
            "movement_summary": summary_for(category, from_club, to_club),
            "from_club": from_club,
            "to_club": to_club,
            "period": period,
            "official_url": value(row, "公式URL"),
            "source_title": "",
            "verification_status": "確認済" if value(row, "公式URL") else "要再確認",
            "checked_at": TODAY,
            "smartphone_note": build_note(category, value(row, "発表日"), from_club, to_club, period, value(row, "メモ")),
            "internal_notes": value(row, "メモ"),
        })
    return players, events


def extract_ladies(path: Path):
    wb = load_workbook(path)
    players = []
    ws_players = wb["選手一覧"]
    for row in sheet_rows(ws_players, 1):
        name = value(row, "氏名")
        if not name:
            continue
        players.append({
            "player_id": f"ladies_2026_{value(row, '背番号', '未定')}_{slug(name)}",
            "team": "ladies",
            "season": "2026/27",
            "number": value(row, "背番号"),
            "position": value(row, "ポジション"),
            "name": name,
            "player_page_url": value(row, "選手ページURL"),
            "base_status": value(row, "2026/27ステータス") or "確認中",
            "notes": value(row, "期間/備考"),
        })
    by_name = {p["name"]: p for p in players}
    events = []
    ws_events = wb["オフ動向"]
    for i, row in enumerate(sheet_rows(ws_events, 1), start=1):
        name = value(row, "氏名")
        if not name:
            continue
        p = by_name.get(name)
        if not p:
            continue
        category = normalize_category(value(row, "動向種別"))
        from_to = value(row, "移籍先/元")
        from_club = from_to if category == "レンタル復帰" else ""
        to_club = from_to if category != "レンタル復帰" else ""
        events.append({
            "event_id": f"ladies_{value(row, '発表日')}_{slug(name)}_{i}",
            "player_id": p["player_id"],
            "announcement_date": value(row, "発表日"),
            "category": category,
            "movement_summary": summary_for(category, from_club, to_club),
            "from_club": from_club,
            "to_club": to_club,
            "period": "",
            "official_url": value(row, "出典URL"),
            "source_title": "",
            "verification_status": "確認済" if value(row, "出典URL") else "要再確認",
            "checked_at": TODAY,
            "smartphone_note": build_note(category, value(row, "発表日"), from_club, to_club, "", ""),
            "internal_notes": value(row, "内容"),
        })
    return players, events


def summary_for(category: str, from_club: str, to_club: str) -> str:
    if category == "契約更新":
        return "契約更新"
    if category == "完全移籍退団":
        return "完全移籍"
    if category == "完全移籍加入":
        return "完全加入"
    if category == "レンタル移籍延長":
        return "レンタル延長"
    if category == "レンタル復帰":
        return "復帰"
    if category == "契約満了":
        return "契約満了"
    if category == "現役引退":
        return "現役引退"
    if category == "レンタル中":
        return "レンタル中"
    return category or "未発表"


def clear_and_create(wb, names):
    for name in list(wb.sheetnames):
        del wb[name]
    for name in names:
        wb.create_sheet(name)


def write_sheet(ws, headers, rows, table_name=None):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)
    if table_name and ws.max_row >= 2:
        ref = f"A1:{ws.cell(1, ws.max_column).coordinate[0]}{ws.max_row}"
        # coordinate[0] is insufficient after Z; current sheets are below Z columns.
        ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {1: 18, 2: 14, 3: 14, 4: 14, 5: 18, 6: 18, 7: 28, 8: 18, 9: 18, 10: 32, 11: 30, 12: 18, 13: 18, 14: 24}
    for idx, width in widths.items():
        if idx <= ws.max_column:
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width


def add_conditional_formatting(ws, category_col_letter="D", max_row=200):
    rng = f"A2:{ws.cell(max_row, ws.max_column).coordinate}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(${category_col_letter}2="契約更新")'], fill=COLOR_FILLS["緑"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(${category_col_letter}2="完全加入",${category_col_letter}2="レンタル加入",${category_col_letter}2="レンタル延長",${category_col_letter}2="復帰")'], fill=COLOR_FILLS["青"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(${category_col_letter}2="完全移籍",${category_col_letter}2="契約満了",${category_col_letter}2="現役引退",${category_col_letter}2="引退")'], fill=COLOR_FILLS["橙"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(${category_col_letter}2="未発表",${category_col_letter}2="公式発表未確認",${category_col_letter}2="レンタル中")'], fill=COLOR_FILLS["黄"]))


def rebuild(source_path: Path, output_path: Path, team: str):
    players, events = extract_standard(source_path) if {"選手マスタ", "オフ動向_入力"}.issubset(set(load_workbook(source_path, read_only=True).sheetnames)) else (extract_men(source_path) if team == "men" else extract_ladies(source_path))
    wb = load_workbook(source_path)
    clear_and_create(wb, ["スマホ表示", "オフ動向_入力", "選手マスタ", "区分マスタ", "確認ログ", "メモ"])
    ws_master = wb["選手マスタ"]
    player_headers = ["player_id", "team", "season", "number", "position", "name", "player_page_url", "base_status", "notes"]
    write_sheet(ws_master, player_headers, players, f"{team}_players")

    ws_events = wb["オフ動向_入力"]
    event_headers = ["event_id", "player_id", "announcement_date", "category", "movement_summary", "from_club", "to_club", "period", "official_url", "source_title", "verification_status", "checked_at", "smartphone_note", "internal_notes"]
    write_sheet(ws_events, event_headers, events, f"{team}_events")
    dv = DataValidation(type="list", formula1="=区分マスタ!$A$2:$A$13", allow_blank=False)
    ws_events.add_data_validation(dv)
    dv.add(f"D2:D{max(200, ws_events.max_row)}")

    ws_cat = wb["区分マスタ"]
    write_sheet(ws_cat, ["category", "display_label", "priority", "color", "smartphone_note_template"], [dict(zip(["category", "display_label", "priority", "color", "smartphone_note_template"], r)) for r in CATEGORY_ROWS], f"{team}_categories")

    ws_log = wb["確認ログ"]
    write_sheet(ws_log, ["checked_at", "team", "page_url", "result", "operator_note"], [
        {"checked_at": TODAY, "team": team, "page_url": "https://www.albirex.co.jp/news/top_team/" if team == "men" else "https://albirex-niigata-ladies.com/news/", "result": "既存データを新仕様へ移行", "operator_note": "公式発表イベントと選手マスタへ分離"}
    ], f"{team}_check_log")

    ws_mobile = wb["スマホ表示"]
    mobile_headers = ["背番号", "ポジション", "氏名", "動向", "その他備考"]
    ws_mobile.append(mobile_headers)
    for i in range(2, len(players) + 2):
        ws_mobile.cell(i, 1).value = f'=IFERROR(INDEX(選手マスタ!$D:$D,ROW()),"")'
        ws_mobile.cell(i, 2).value = f'=IFERROR(INDEX(選手マスタ!$E:$E,ROW()),"")'
        ws_mobile.cell(i, 3).value = f'=IFERROR(INDEX(選手マスタ!$F:$F,ROW()),"")'
        ws_mobile.cell(i, 4).value = f'=IFERROR(XLOOKUP(INDEX(選手マスタ!$A:$A,ROW()),オフ動向_入力!$B:$B,オフ動向_入力!$E:$E,"未発表",0,-1),"未発表")'
        ws_mobile.cell(i, 5).value = f'=IFERROR(XLOOKUP(INDEX(選手マスタ!$A:$A,ROW()),オフ動向_入力!$B:$B,オフ動向_入力!$M:$M,"公式発表未確認",0,-1),"公式発表未確認")'
    style_sheet(ws_mobile)
    ws_mobile.column_dimensions["A"].width = 10
    ws_mobile.column_dimensions["B"].width = 10
    ws_mobile.column_dimensions["C"].width = 18
    ws_mobile.column_dimensions["D"].width = 18
    ws_mobile.column_dimensions["E"].width = 32
    add_conditional_formatting(ws_mobile, "D", max(200, len(players) + 20))

    ws_memo = wb["メモ"]
    memo_rows = [
        {"項目": "対象", "内容": "アルビレックス新潟 男子トップチーム" if team == "men" else "アルビレックス新潟レディース"},
        {"項目": "改修日", "内容": TODAY},
        {"項目": "正本シート", "内容": "選手マスタ、オフ動向_入力"},
        {"項目": "スマホ表示", "内容": "背番号／ポジション／氏名／動向／その他備考の5列のみ。手入力禁止。"},
        {"項目": "色分け", "内容": "条件付き書式で自動適用。手作業でセル色を変更しない。"},
        {"項目": "更新手順", "内容": "公式発表1件につきオフ動向_入力へ1行追加。新規選手は選手マスタにも追加。"},
        {"項目": "仕様書", "内容": "docs/offseason-update-spec.md"},
    ]
    write_sheet(ws_memo, ["項目", "内容"], memo_rows, f"{team}_memo")
    ws_memo.column_dimensions["A"].width = 18
    ws_memo.column_dimensions["B"].width = 80

    # Basic workbook settings.
    wb.active = 0
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"rebuilt {output_path}: players={len(players)} events={len(events)}")


def main():
    rebuild(
        ROOT / "albirex_niigata_men_2026_27_offseason_print_mobile.xlsx",
        EXPORTS / "albirex_niigata_men_2026_27_offseason_print_mobile.xlsx",
        "men",
    )
    rebuild(
        ROOT / "albirex_niigata_ladies_2026_27_offseason_print_mobile.xlsx",
        EXPORTS / "albirex_niigata_ladies_2026_27_offseason_print_mobile.xlsx",
        "ladies",
    )


if __name__ == "__main__":
    main()
